from django.shortcuts import render
from django.http import HttpResponse
# модули для работы с xls конвертером
import openpyxl, calendar, datetime
import sys
# модуль для работы получения ссылки на созданный файл
import mimetypes


# Переводит вводимую дату в последний день введенного месяца
def eop(date):
    last_day = calendar.monthrange(date.year, date.month)[1]
    return date.replace(day=last_day)

# Проверяет тип введенной даты и переводит текст "дд.мм.гггг" в тип datetime
def check_date(cell):
    if not isinstance(cell.value, datetime.date):
        x = [int(i) for i in cell.value.split('.')]
        return datetime.date(day=x[0], month=x[1], year=x[2])
    else:
        return cell.value

# Проверяет, есть ли в ячейке гиперссылка и достает её или возвращает обычное значение
def check_link(cell):
    try:
        if cell.hyperlink.target:
            return cell.hyperlink.target
    except:
        return cell.value


def index(request):
    if request.method == 'GET':
        return render(request, 'downloadfiles_app/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        wb_in = openpyxl.load_workbook(excel_file)
        sheet_in = wb_in.active

        # Определяем размер заполненной входящей таблицы, для задания границ обработки информации
        max_col = sheet_in.max_column
        max_row = sheet_in.max_row

        # Формируем список полей файла и словарь
        titles = [cell.value for cell in sheet_in['1']]
        print(titles)
        titles_nums = [i for i in range(max_col)]
        titles_dic = dict(zip(titles, titles_nums))
        print(titles_dic)

        # Проверяем наличие обязательных полей. Останавливаем и выдаем сообщение, если не хватает.
        titles_must = {'Начало периода', 'Конец периода', 'Фото', 'ТТ'}
        titles_asis = set(titles)
        if not titles_must.issubset(titles_asis):
            print('Error: Отсутствуют обязательные поля. Добавьте в файл колонки с именами:', titles_must - titles_asis)
            sys.exit()

        # Узнаем номер колонки нужного поля: "начало" и "конец" РК
        start_day_indx = titles_dic['Начало периода']
        end_day_indx = titles_dic['Конец периода']

        # Считываем строки со значениями и обрабатываю по ходу
        result = [titles]
        for row in range(2, max_row + 1):
            side = []
            for col in range(1, max_col + 1):
                cell = sheet_in.cell(row=row, column=col)
                if col == titles_dic['Фото'] + 1 or col == titles_dic['ТТ'] + 1:
                    cell.value = check_link(cell)
                if col == titles_dic['Начало периода'] + 1 or col == titles_dic['Конец периода'] + 1:
                    cell.value = check_date(cell)
                side.append(cell.value)
            print(side)

            # Проверям кол-во месяцев в периоде и создаю такое кол-во строк в выходном списке
            start_month = side[start_day_indx].month
            end_month = side[end_day_indx].month
            months = end_month - start_month + 1

            # Создаем нужное кол-во копий стороны в результирующем файле и разбиваем даты
            if months == 1:
                result.append(side.copy())
            elif months >= 2:
                # У первого месяца оставляем введенную дату Начала РК и задаем последнюю дату месяца Конца РК
                result.append(side.copy())
                result[-1][end_day_indx] = eop(result[-1][start_day_indx])
                # Промежуточные месяцы имеют период: с первого по последнее число месяца
                for i in range(1, months - 1):
                    result.append(side.copy())
                    result[-1][start_day_indx] = side[start_day_indx].replace(day=1, month=(start_month + i))
                    result[-1][end_day_indx] = eop(result[-1][start_day_indx])
                # У последнего месяца оставляем введенную дату Конца РК и задаем первое число месяца Начала РК
                result.append(side.copy())
                result[-1][start_day_indx] = result[-1][end_day_indx].replace(day=1)

        # Проверяем результат
        print()
        [print(i) for i in result]

        # Создаем лист для исходящего файла
        wb_out = openpyxl.Workbook()
        sheet_out = wb_out.active

        # Записываем в выходной файл
        for row in range(1, len(result) + 1):
            for col in range(1, len(titles) + 1):
                c = sheet_out.cell(row=row, column=col)
                c.value = result[row - 1][col - 1]

        # Автоматическая ширина колонок (исключая ошибки при пустых колонках)
        for col in sheet_out.columns:
            width = 8
            max_width = 50
            ltr = col[0].column_letter  # get the column letter
            for cell in col[1:]:  # ignoring labels length
                try:
                    if len(str(cell.value)) > max_width:
                        width = max_width
                    elif len(str(cell.value)) > width:
                        width = len(str(cell.value))
                except:
                    pass
            adjusted_width = (width + 2) * 1.2
            sheet_out.column_dimensions[ltr].width = adjusted_width

        # Конвертируем даты в формат 'ДД.ММ.ГГГГ'
        for col in sheet_out.columns:
            for cell in col:
                if isinstance((cell.value), datetime.date):
                    cell.number_format = 'DD.MM.YYYY'

        wb_out.save("media/media/result.xlsx")
        return render(request, 'downloadfiles_app/result.html')


def result(request):
    return render(request, 'downloadfiles_app/result.html')


def download_file(request):
    with open('media/media/result.xlsx', 'rb') as f:
        response = HttpResponse(f.read(), content_type="application/ms-excel")
        response['Content-Disposition'] = 'attachment; filename=result.xlsx'
        return response
